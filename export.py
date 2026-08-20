"""Excel / PDF export for a sales summary period.

Both builders take a summary dict — either db.get_month_summary() or
db.get_year_summary() — plus the matching raw order list, and return raw bytes
ready to hand back as a Flask file download. The two summaries differ only in
how the period is named and broken down, so _period() below flattens that
difference and the builders themselves stay period-agnostic.
"""

from __future__ import annotations

import io


def _period(summary: dict) -> dict:
    """Flatten a month-or-year summary into the labels/rows both exports need."""
    if "by_month" in summary:
        return {
            "label": summary["year"],
            "best": summary.get("best_month"),
            "best_label": lambda b: b["month"],
            "rows_title": "Orders per month",
            "rows_head": "Month",
            "rows": [(r["month"], r["orders"], r["revenue"]) for r in summary["by_month"]],
        }
    return {
        "label": summary["month"],
        "best": summary.get("best_day"),
        "best_label": lambda b: b["date"],
        "rows_title": "Orders per day",
        "rows_head": "Day",
        "rows": [(r["date"], r["orders"], r["revenue"]) for r in summary["by_day"]],
    }


def build_excel(summary: dict, orders: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    per = _period(summary)
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    bold = Font(bold=True)
    ws.append([f"Sales summary — {per['label']}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Orders", summary["orders"]])
    ws.append(["Revenue", summary["revenue"]])
    ws.append(["Glasses sold", summary.get("drinks_qty", 0)])
    ws.append(["Desserts sold", summary.get("desserts_qty", 0)])
    if "trading_days" in summary:
        ws.append(["Days open", summary["trading_days"]])
        ws.append(["Average per day open", summary["avg_per_trading_day"]])
    for pm, amt in summary["by_payment"].items():
        ws.append([pm, amt])
    if per["best"]:
        ws.append([f"Busiest {per['rows_head'].lower()}", per["best_label"](per["best"]),
                   per["best"]["orders"], "orders"])
    ws.append([])
    ws.append(["Items sold"])
    ws[f"A{ws.max_row}"].font = bold
    ws.append(["Item", "Temperature", "Quantity sold", "Revenue"])
    for c in ws[ws.max_row]:
        c.font = bold
    for it in summary["top_items"]:
        ws.append([it["name"], it["temperature"], it["quantity"], it["revenue"]])
    ws.append([])
    ws.append([per["rows_head"], "Orders", "Revenue"])
    for c in ws[ws.max_row]:
        c.font = bold
    for label, orders_n, revenue in per["rows"]:
        ws.append([label, orders_n, revenue])
    for col in "ABCD":
        ws.column_dimensions[col].width = 20

    ws2 = wb.create_sheet("Orders")
    ws2.append(["Order ID", "Date", "Time", "Payment", "Item", "Temperature",
                "Qty", "Unit price", "Line total", "Order total", "Comment"])
    for c in ws2[1]:
        c.font = bold
    for o in orders:
        time = (o["created_at"] or "")[11:16]
        if not o["lines"]:
            ws2.append([o["id"], o["sale_date"], time, o["payment_method"],
                        "", "", "", "", "", o["total"], o["comment"]])
            continue
        for ln in o["lines"]:
            ws2.append([o["id"], o["sale_date"], time, o["payment_method"],
                        ln["name"], ln["temperature"], ln["quantity"],
                        ln["unit_price"], ln["line_total"], o["total"], o["comment"]])
    for col in "ABCDEFGHIJK":
        ws2.column_dimensions[col].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf(summary: dict, orders: list[dict]) -> bytes:
    from fpdf import FPDF

    per = _period(summary)
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, f"Sales summary - {per['label']}", ln=True)

    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 8, f"Orders: {summary['orders']}    Revenue: THB {summary['revenue']:.2f}", ln=True)
    pdf.cell(0, 7, f"Glasses sold: {summary.get('drinks_qty', 0):g}    "
                   f"Desserts sold: {summary.get('desserts_qty', 0):g}", ln=True)
    if "trading_days" in summary:
        pdf.cell(0, 7, f"Days open: {summary['trading_days']}    "
                       f"Average per day open: THB {summary['avg_per_trading_day']:.2f}", ln=True)
    for pm, amt in summary["by_payment"].items():
        pdf.cell(0, 7, f"  {pm}: THB {amt:.2f}", ln=True)
    if per["best"]:
        bd = per["best"]
        pdf.cell(0, 7, f"Busiest {per['rows_head'].lower()}: {per['best_label'](bd)} "
                       f"({bd['orders']} orders, THB {bd['revenue']:.2f})", ln=True)

    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, "Items sold", ln=True)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(90, 7, "Item", border=1)
    pdf.cell(30, 7, "Temp", border=1)
    pdf.cell(30, 7, "Qty", border=1)
    pdf.cell(30, 7, "Revenue", border=1, ln=True)
    pdf.set_font("Helvetica", "", 10)
    for it in summary["top_items"]:
        pdf.cell(90, 7, str(it["name"])[:45], border=1)
        pdf.cell(30, 7, str(it["temperature"] or "-"), border=1)
        pdf.cell(30, 7, f"{it['quantity']:g}", border=1)
        pdf.cell(30, 7, f"{it['revenue']:.2f}", border=1, ln=True)

    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, per["rows_title"], ln=True)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(60, 7, per["rows_head"], border=1)
    pdf.cell(40, 7, "Orders", border=1)
    pdf.cell(40, 7, "Revenue", border=1, ln=True)
    pdf.set_font("Helvetica", "", 10)
    for label, orders_n, revenue in per["rows"]:
        if orders_n == 0 and revenue == 0:
            continue
        pdf.cell(60, 7, label, border=1)
        pdf.cell(40, 7, str(orders_n), border=1)
        pdf.cell(40, 7, f"{revenue:.2f}", border=1, ln=True)

    return bytes(pdf.output())
